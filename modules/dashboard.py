import streamlit as st
import pandas as pd
import plotly.express as px
import io
from database.connection import run_query
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BRASILIA = ZoneInfo("America/Sao_Paulo")
PREFIXO_FECHAMENTO = "Informar fechamento de período"

HEADER_FILL = PatternFill("solid", fgColor="041747")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=11, color="000000")
DATA_ALIGN = Alignment(horizontal="left", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F2F4F8")
BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD")
)

@st.cache_data(ttl=30)
def carregar_chamados():
    return run_query("""
        SELECT protocolo, setor, empresa, tipo_inconsistencia,
               prioridade, status, aberto_em, atendido_em, resolvido_em
        FROM chamados ORDER BY aberto_em DESC
    """, fetch=True)

@st.cache_data(ttl=30)
def carregar_chamados_completo():
    return run_query("""
        SELECT protocolo, setor, empresa, tipo_inconsistencia,
               prioridade, nf_retorna, nome_parceiro, numero_nota,
               tipo_nota, data_entrada, data_saida, data_negociacao,
               valor, observacao, status, aberto_em, atendido_em,
               resolvido_em, resolucao
        FROM chamados ORDER BY aberto_em DESC
    """, fetch=True)

def estilo_cabecalho(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

def estilo_dados(ws, row_start, row_end, num_cols):
    for row in range(row_start, row_end + 1):
        fill = ALT_FILL if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.fill = fill
            cell.border = BORDER

def ajustar_colunas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

def inserir_cabecalho_relatorio(ws, titulo):
    try:
        img = XLImage("assets/LOGO-GRUPO-LLE-COR-OFICIAL-PRINCIPAL.png")
        img.width = 160
        img.height = 55
        ws.add_image(img, "A1")
    except:
        pass
    ws.row_dimensions[1].height = 35
    ws["E1"] = titulo
    ws["E1"].font = Font(name="Calibri", bold=True, size=14, color="041747")
    ws["E2"] = "Grupo LLE"
    ws["E2"].font = Font(name="Calibri", size=11, color="0071FE", bold=True)
    ws["E3"] = f"Gerado em: {datetime.now(BRASILIA).strftime('%d/%m/%Y às %H:%M')}"
    ws["E3"].font = Font(name="Calibri", size=10, color="999999")

def tela_dashboard():
    st.title("📊 Dashboard")
    st.markdown("---")

    rows = carregar_chamados()
    if not rows:
        st.info("Nenhum chamado registrado ainda.")
        return

    df = pd.DataFrame(rows, columns=[
        "protocolo","setor","empresa","tipo",
        "prioridade","status","aberto_em","atendido_em","resolvido_em"
    ])
    df["aberto_em"] = pd.to_datetime(df["aberto_em"])
    df["resolvido_em"] = pd.to_datetime(df["resolvido_em"])
    df["tempo_resolucao"] = (df["resolvido_em"] - df["aberto_em"]).dt.total_seconds() / 3600

    st.markdown("#### 🔎 Filtros")
    c1, c2, c3 = st.columns(3)
    filtro_status = c1.multiselect("Status", df["status"].unique().tolist(), default=df["status"].unique().tolist())
    filtro_empresa = c2.multiselect("Empresa", df["empresa"].unique().tolist(), default=df["empresa"].unique().tolist())
    filtro_setor = c3.multiselect("Setor", df["setor"].unique().tolist(), default=df["setor"].unique().tolist())

    # Filtro por data (base: Abertura ou Resolução) + período
    d1, d2, d3 = st.columns(3)
    campo_label = d1.selectbox("Filtrar data por", ["Abertura", "Resolução"])
    col_data = "aberto_em" if campo_label == "Abertura" else "resolvido_em"
    data_min = df["aberto_em"].min().date()
    data_max = df["aberto_em"].max().date()
    data_ini = d2.date_input("De", value=data_min, key="dash_data_ini")
    data_fim = d3.date_input("Até", value=data_max, key="dash_data_fim")

    mask_data = df[col_data].dt.date.between(data_ini, data_fim)

    df_f = df[
        df["status"].isin(filtro_status) &
        df["empresa"].isin(filtro_empresa) &
        df["setor"].isin(filtro_setor) &
        mask_data
    ]

    st.markdown("---")
    st.markdown("#### 📈 Indicadores")
    k1, k2, k3, k4, k5 = st.columns(5)
    total = len(df_f)
    abertos = len(df_f[df_f["status"] == "Aberto"])
    em_andamento = len(df_f[df_f["status"] == "Em andamento"])
    resolvidos = len(df_f[df_f["status"] == "Resolvido"])
    tempo_medio = df_f[df_f["tempo_resolucao"].notna()]["tempo_resolucao"].mean()
    k1.metric("Total", total)
    k2.metric("🔴 Abertos", abertos)
    k3.metric("🟡 Em andamento", em_andamento)
    k4.metric("🟢 Resolvidos", resolvidos)
    k5.metric("⏱️ Tempo médio", f"{tempo_medio:.1f}h" if not pd.isna(tempo_medio) else "—")

    st.markdown("---")
    ca, cb = st.columns(2)
    with ca:
        st.markdown("##### Chamados por Tipo")
        df_t = df_f.groupby("tipo").size().reset_index(name="qtd").sort_values("qtd", ascending=False)
        fig1 = px.bar(df_t, x="qtd", y="tipo", orientation="h", color="qtd", color_continuous_scale="Blues", labels={"qtd":"Qtd","tipo":""})
        fig1.update_layout(showlegend=False, coloraxis_showscale=False, margin=dict(l=0,r=0,t=0,b=0), height=300)
        st.plotly_chart(fig1, use_container_width=True)
    with cb:
        st.markdown("##### Chamados por Setor")
        df_s = df_f.groupby("setor").size().reset_index(name="qtd").sort_values("qtd", ascending=False)
        fig2 = px.bar(df_s, x="qtd", y="setor", orientation="h", color="qtd", color_continuous_scale="Greens", labels={"qtd":"Qtd","setor":""})
        fig2.update_layout(showlegend=False, coloraxis_showscale=False, margin=dict(l=0,r=0,t=0,b=0), height=300)
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("##### 🍕 Tipos de chamado por Setor")
    setores_unicos = sorted([s for s in df_f["setor"].dropna().unique().tolist() if s])
    if not setores_unicos:
        st.info("Sem dados para o período selecionado.")
    else:
        n_cols = 2
        for inicio in range(0, len(setores_unicos), n_cols):
            grupo = setores_unicos[inicio:inicio+n_cols]
            cols = st.columns(n_cols)
            for j, setor in enumerate(grupo):
                with cols[j]:
                    df_pie = df_f[df_f["setor"] == setor].groupby("tipo").size().reset_index(name="qtd")
                    fig = px.pie(df_pie, names="tipo", values="qtd")
                    fig.update_traces(textposition="inside", textinfo="percent")
                    fig.update_layout(
                        title=dict(text=setor, font=dict(size=14, color="#041747")),
                        margin=dict(l=0, r=0, t=40, b=0), height=320,
                        legend=dict(orientation="h", y=-0.1, font=dict(size=9))
                    )
                    st.plotly_chart(fig, use_container_width=True, key=f"pie_setor_{inicio+j}")

    st.markdown("---")
    st.markdown("##### 📈 Evolução Mensal")
    df_evo = df_f.copy()
    df_evo["mes"] = df_evo["aberto_em"].dt.to_period("M").astype(str)
    df_m = df_evo.groupby("mes").size().reset_index(name="qtd").sort_values("mes")
    fig5 = px.line(df_m, x="mes", y="qtd", markers=True, labels={"mes":"Mês","qtd":"Chamados"})
    fig5.update_layout(margin=dict(l=0,r=0,t=0,b=0), height=250)
    st.plotly_chart(fig5, use_container_width=True)

    # === Registro de entregas de fechamento de período ===
    st.markdown("---")
    st.markdown("##### 🗂️ Entregas de Fechamento de Período")
    st.caption("Filtra pela data da entrega (data de abertura) e pelos setores selecionados.")
    mask_entrega_data = df["aberto_em"].dt.date.between(data_ini, data_fim)
    df_entregas = df[
        df["setor"].isin(filtro_setor) &
        mask_entrega_data &
        df["tipo"].astype(str).str.startswith(PREFIXO_FECHAMENTO)
    ].copy()

    if df_entregas.empty:
        st.info("Nenhuma entrega de fechamento de período no período/setores selecionados.")
    else:
        df_entregas["Período"] = (
            df_entregas["tipo"].astype(str)
            .str.replace(f"{PREFIXO_FECHAMENTO} - ", "", regex=False)
            .str.replace(PREFIXO_FECHAMENTO, "—", regex=False)
        )
        df_entregas["Data/Hora"] = df_entregas["aberto_em"].dt.strftime("%d/%m/%Y %H:%M")
        tabela_entregas = (
            df_entregas[["setor","Período","Data/Hora","protocolo"]]
            .rename(columns={"setor":"Setor","protocolo":"Protocolo"})
            .sort_values("Data/Hora", ascending=False)
        )
        st.dataframe(tabela_entregas, use_container_width=True, hide_index=True)
        st.caption(f"Total de entregas: {len(tabela_entregas)}")

    st.markdown("---")
    st.markdown("##### 📥 Exportar dados")
    todos = carregar_chamados_completo()
    df_export = pd.DataFrame(todos, columns=[
        "Protocolo","Setor","Empresa","Abertura de Período / Descontabilização",
        "Prioridade","NF Retorna","Parceiro","Número Nota",
        "Tipo Nota","Data Entrada","Data Saída","Data Negociação",
        "Valor","Observação","Status","Aberto Em","Atendido Em",
        "Resolvido Em","Resolução"
    ])
    df_kpi = pd.DataFrame({"Indicador":["Total","Abertos","Em Andamento","Resolvidos","Tempo Médio (h)"],
                            "Valor":[total,abertos,em_andamento,resolvidos,f"{tempo_medio:.1f}" if not pd.isna(tempo_medio) else "—"]})
    df_te = df_f.groupby("tipo").size().reset_index(name="Quantidade").sort_values("Quantidade",ascending=False).rename(columns={"tipo":"Tipo"})
    df_se = df_f.groupby("setor").size().reset_index(name="Quantidade").sort_values("Quantidade",ascending=False).rename(columns={"setor":"Setor"})
    df_ee = df_f.groupby("empresa").size().reset_index(name="Quantidade").rename(columns={"empresa":"Empresa"})
    df_ste = df_f.groupby("status").size().reset_index(name="Quantidade").rename(columns={"status":"Status"})

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Chamados", startrow=6)
        ws1 = writer.sheets["Chamados"]
        inserir_cabecalho_relatorio(ws1, "ROC — Registro de Ocorrências Contábeis")
        estilo_cabecalho(ws1, 7, len(df_export.columns))
        estilo_dados(ws1, 8, 7+len(df_export), len(df_export.columns))
        ws1.row_dimensions[7].height = 30
        ajustar_colunas(ws1)
        ws1.freeze_panes = "A8"

        ws2 = writer.book.create_sheet("Dashboard")
        writer.sheets["Dashboard"] = ws2
        inserir_cabecalho_relatorio(ws2, "ROC — Dashboard Operacional")
        secoes = [("📊 KPIs",df_kpi,"041747",False),("📌 Por Tipo",df_te,"041747",True),
                  ("🏢 Por Setor",df_se,"0F8C3B",True),("🏭 Por Empresa",df_ee,"0071FE",True),
                  ("🔘 Por Status",df_ste,"FAC318",True)]
        linha = 6
        for titulo, df_sec, cor_hex, usa_alt in secoes:
            ws2.cell(row=linha,column=1,value=titulo).font = Font(name="Calibri",bold=True,size=12,color="041747")
            ws2.cell(row=linha,column=1).fill = PatternFill("solid",fgColor="F0F4FF")
            ws2.row_dimensions[linha].height = 22
            linha += 1
            df_sec.to_excel(writer, index=False, sheet_name="Dashboard", startrow=linha-1)
            font_cor = "041747" if cor_hex == "FAC318" else "FFFFFF"
            for col_num in range(1, len(df_sec.columns)+1):
                cell = ws2.cell(row=linha, column=col_num)
                cell.fill = PatternFill("solid", fgColor=cor_hex)
                cell.font = Font(name="Calibri", bold=True, color=font_cor, size=11)
                cell.alignment = HEADER_ALIGN
                cell.border = BORDER
            ws2.row_dimensions[linha].height = 25
            for r in range(linha+1, linha+len(df_sec)+1):
                alt = ALT_FILL if (r%2==0 and usa_alt) else PatternFill("solid",fgColor="FFFFFF")
                for col_num in range(1, len(df_sec.columns)+1):
                    cell = ws2.cell(row=r, column=col_num)
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(horizontal="center",vertical="center")
                    cell.fill = alt
                    cell.border = BORDER
            linha += len(df_sec) + 3
        ajustar_colunas(ws2)

    buffer.seek(0)
    st.download_button(
        label="📥 Baixar Excel completo",
        data=buffer,
        file_name=f"ROC_{datetime.now(BRASILIA).strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
